import pandas as pd 
from openpyxl import load_workbook
from tkinter import *
from tkinter import messagebox, simpledialog, Toplevel, Text, WORD, END, Button, Canvas, Scrollbar
from tkinter import ttk
import tkinter as tk
import os
import subprocess
from openpyxl import Workbook, load_workbook
from matplotlib import pyplot as plt
from matplotlib.dates import DateFormatter
from PIL import Image, ImageTk

# Lista de archivos Excel disponibles
archivos_excel = []
Mant = 1

# Ruta del archivo Excel
archivo_excel = 'Motores.xlsx'

# Verifica si el archivo existe
if not os.path.exists(archivo_excel):
    # Si no existe, crea un nuevo archivo Excel
    libro = Workbook()
    hoja = libro.active
    hoja.title = 'Hoja1'  # Le asignamos un nombre a la hoja
    libro.save(archivo_excel)
    print(f"El archivo {archivo_excel} ha sido creado.")
else:
    print(f"El archivo {archivo_excel} ya existe.")
    

def crear_excel(nombre_archivo):
    # Crear un archivo Excel nuevo si no existe
    if not os.path.isfile(nombre_archivo):
        df = pd.DataFrame()
        df.to_excel(nombre_archivo, index=False, sheet_name='Hoja1')
        actualizar_combobox_archivos()

def guardar_motor_en_excel(nuevo_motor, hoja_seleccionada, nombre_archivo):
    book = load_workbook(nombre_archivo)
    if hoja_seleccionada not in book.sheetnames:
        messagebox.showerror("Error", "La hoja seleccionada no existe.")
        return

    hoja = book[hoja_seleccionada]

    if hoja.max_row == 1:  # Si hay solo las cabeceras
        last_row = 1  # La primera fila de datos será la primera
    else:
        last_row = hoja.max_row + 2  # Dejar una fila vacía entre las listas

    for index, (key, value) in enumerate(nuevo_motor.items()):
        hoja.cell(row=last_row + index, column=1, value=key)  # Escribir el nombre del campo
        hoja.cell(row=last_row + index, column=2, value=value)  # Escribir solo el valor del campo

    book.save(nombre_archivo)
    messagebox.showinfo("Éxito", "Motor guardado exitosamente.")

def crear_hoja(nombre_archivo):
    nombre_hoja = simpledialog.askstring("Crear hoja", "Ingrese el nombre de la nueva hoja:")
    if nombre_hoja:
        book = load_workbook(nombre_archivo)
        if nombre_hoja not in book.sheetnames:
            book.create_sheet(title=nombre_hoja)
            book.save(nombre_archivo)
            messagebox.showinfo("Éxito", "Hoja creada exitosamente.")
            actualizar_combobox_hojas()
        else:
            messagebox.showerror("Error", "La hoja ya existe.")

def guardar_nuevo_motor():
    nuevo_motor = {
        'AREA': Area_entry.get(),
        'CÓDIGO DE EQUIPO': CODIGO_entry.get(),
        'TIPO DE MOTOR': TIPO_entry.get(),
        'DATOS DE MOTOR': DATOS_entry.get(),
        'MARCA': MARCA_entry.get(),
        'FUNCIÓN': FUNCION_entry.get(),
        'HP/KW': POTENCIA_entry.get(),
        'VOLTAJE': VOLTAJE_entry.get(),
        'VOLTAJE FUENTE': VOLTAJE_FUENTE_entry.get(),
        'AMPERAJE': AMPERAJE_entry.get(),
        'COJINETE FRONTAL': COJINETE_FRONTAL_entry.get(),
        'COJINETE TRASERO': COJINETE_TRASERO_entry.get(),
        'MANTENIMIENTO': MANTENIMIENTO_entry.get(),
        'RPM': RPM_entry.get(),
        'HERTZ': HERTZ_entry.get(),
        'OTROS': OTROS_entry.get()
    }
    
    # Aquí puedes añadir la lógica para guardar el motor en el archivo Excel o base de datos.
    hoja_seleccionada = hoja_combobox.get()
    nombre_archivo = archivo_combobox.get()
    guardar_motor_en_excel(nuevo_motor, hoja_seleccionada, nombre_archivo)

    # Limpiar los campos después de guardar
    Area_entry.delete(0, END)
    CODIGO_entry.delete(0, END)
    TIPO_entry.delete(0, END)
    DATOS_entry.delete(0, END)
    MARCA_entry.delete(0, END)
    FUNCION_entry.delete(0, END)
    POTENCIA_entry.delete(0, END)
    VOLTAJE_entry.delete(0, END)
    VOLTAJE_FUENTE_entry.delete(0, END)
    AMPERAJE_entry.delete(0, END)
    COJINETE_FRONTAL_entry.delete(0, END)
    COJINETE_TRASERO_entry.delete(0, END)
    MANTENIMIENTO_entry.delete(0, END)
    RPM_entry.delete(0, END)
    HERTZ_entry.delete(0, END)
    OTROS_entry.delete(0, END)
    
    # Opcional: Mostrar un mensaje de éxit
    messagebox.showinfo("Éxito", "Motor guardado y campos vacíos")


def editar_motor():
    hoja_seleccionada = hoja_combobox.get()
    nombre_archivo = archivo_combobox.get()

    book = load_workbook(nombre_archivo)

    if hoja_seleccionada not in book.sheetnames:
        messagebox.showerror("Error", "La hoja seleccionada no existe.")
        return

    hoja = book[hoja_seleccionada]

    motores = []
    # Suponemos que los nombres de los motores están en la segunda fila de cada bloque de información
    for row in range(1, hoja.max_row + 1):
        campo = hoja.cell(row=row, column=1).value
        if campo == "AREA":
            motor_nombre = hoja.cell(row=row, column=2).value
            if motor_nombre:
                motores.append(motor_nombre)

    lista_motores = "\n".join(motores)
    nombre_motor = simpledialog.askstring("Editar Motor", f"Seleccione un motor para editar:\n{lista_motores}\n\nIngrese el area del motor a editar:")

    if not nombre_motor:
        messagebox.showwarning("Advertencia", "Debe ingresar un area de motor.")
        return

    motor_encontrado = False
    for row in range(1, hoja.max_row + 1):
        if hoja.cell(row=row, column=2).value == nombre_motor:
            motor_encontrado = True
            campo_a_editar = simpledialog.askstring("Editar Campo", "Ingrese el campo a editar (AREA, CODIGO, TIPO, DATOS, MARCA, FUNCION, POTENCIA, VOLTAJE, VOLTAJE_F, AMPERAJE, COJINETE FRONTAL, COJINETE TRASERO, MANTENIMIENTO, RPM, HERTZ, OTROS):")
            nuevo_valor = simpledialog.askstring("Nuevo Valor", f"Ingrese el nuevo valor para {campo_a_editar}:")

            if campo_a_editar not in ['AREA', 'CODIGO', 'TIPO', 'DATOS', 'MARCA', 'FUNCION', 'POTENCIA', 'VOLTAJE', 'VOLTAJE_F', 'AMPERAJE', 'COJINETE FRONTAL', 'COJINETE TRASERO', 'MANTENIMIENTO', 'RPM', 'HERTZ', 'OTROS']:
                messagebox.showerror("Error", "Campo no válido.")
                return
            

            campo_index = {
                'AREA': 0,
                'CODIGO': 1,
                'TIPO': 2,
                'DATOS': 3,
                'MARCA': 4,
                'FUNCION': 5,
                'POTENCIA': 6,
                'VOLTAJE': 7,
                'VOLTAJE_F': 8,
                'AMPERAJE': 9,
                'COJINETE FRONTAL': 10,
                'COJINETE TRASERO': 11,
                'MANTENIMIENTO': 12,
                'RPM': 13,
                'HERTZ': 14,
                'OTROS': 15
            }.get(campo_a_editar)

            if campo_index is not None:
                hoja.cell(row=row + campo_index, column=2, value=nuevo_valor)
                break
            else:
                messagebox.showerror("Error", "Campo no válido.")
                return

    if motor_encontrado:
        book.save(nombre_archivo)
        messagebox.showinfo("Éxito", f"Motor '{nombre_motor}' editado exitosamente.")
    else:
        messagebox.showerror("Error", f"Motor '{nombre_motor}' no encontrado en la hoja '{hoja_seleccionada}'.")

def agregar_mantenimiento():
    hoja_seleccionada = hoja_combobox.get()
    nombre_archivo = archivo_combobox.get()

    cojinete = simpledialog.askstring("Agregar Mantenimiento", "Ingrese el nombre del técnico:")
    tipo_mantenimiento = simpledialog.askstring("Agregar Mantenimiento", "Ingrese el Tipo de mantenimiento:")
    fecha = simpledialog.askstring("Agregar Mantenimiento", "Ingrese la Fecha:")

    if not (cojinete and tipo_mantenimiento and fecha):
        messagebox.showerror("Error", "Todos los campos de mantenimiento deben estar completos.")
        return

    # Cargar el archivo Excel
    book = load_workbook(nombre_archivo)
    if hoja_seleccionada not in book.sheetnames:
        messagebox.showerror("Error", "La hoja seleccionada no existe.")
        return

    hoja = book[hoja_seleccionada]

    # Verificar si los encabezados ya existen, si no, crearlos en la primera fila
    if hoja.cell(row=1, column=4).value != "Técnico":
        hoja.cell(row=1, column=4, value="Técnico")
    if hoja.cell(row=1, column=5).value != "Tipo de Mantenimiento":
        hoja.cell(row=1, column=5, value="Tipo de Mantenimiento")
    if hoja.cell(row=1, column=6).value != "Fecha":
        hoja.cell(row=1, column=6, value="Fecha")

    # Encontrar la siguiente fila vacía en las columnas 4, 5 y 6 (Cojinete, Tipo de Mantenimiento, Fecha)
    next_row = 2  # Empezamos desde la segunda fila porque la primera fila son los encabezados
    while (hoja.cell(row=next_row, column=4).value is not None or 
           hoja.cell(row=next_row, column=5).value is not None or 
           hoja.cell(row=next_row, column=6).value is not None): 
        next_row += 1  # Aumentar la fila hasta encontrar una fila vacía en las tres columnas

    # Ingresar los valores en la siguiente fila vacía
    hoja.cell(row=next_row, column=4, value=cojinete)
    hoja.cell(row=next_row, column=5, value=tipo_mantenimiento)
    hoja.cell(row=next_row, column=6, value=fecha)

    # Guardar los cambios en el archivo Excel
    book.save(nombre_archivo)
    messagebox.showinfo("Éxito", "Mantenimiento agregado exitosamente.")
    

def agregar_revision_termica():
    hoja_seleccionada = hoja_combobox.get()
    nombre_archivo = archivo_combobox.get()

    cojinete = simpledialog.askstring("Agregar Mantenimiento", "Ingrese el cojinete:")
    Temperatura = simpledialog.askstring("Agregar Mantenimiento", "Ingrese la temperatura en °C:")
    fecha = simpledialog.askstring("Agregar Mantenimiento", "Ingrese la Fecha:")

    if not (cojinete and Temperatura and fecha):
        messagebox.showerror("Error", "Todos los campos de revisión deben estar completos.")
        return

    # Cargar el archivo Excel
    book = load_workbook(nombre_archivo)
    if hoja_seleccionada not in book.sheetnames:
        messagebox.showerror("Error", "La hoja seleccionada no existe.")
        return

    hoja = book[hoja_seleccionada]

    # Verificar si los encabezados ya existen, si no, crearlos en la primera fila
    if hoja.cell(row=1, column=8).value != "Cojinete":
        hoja.cell(row=1, column=8, value="Cojinete")
    if hoja.cell(row=1, column=9).value != "Temperatura":
        hoja.cell(row=1, column=9, value="Temperatura")
    if hoja.cell(row=1, column=10).value != "Fecha":
        hoja.cell(row=1, column=10, value="Fecha")

    # Encontrar la siguiente fila vacía en las columnas 4, 5 y 6 (Cojinete, Tipo de Mantenimiento, Fecha)
    next_row = 2  # Empezamos desde la segunda fila porque la primera fila son los encabezados
    while (hoja.cell(row=next_row, column=8).value is not None or 
           hoja.cell(row=next_row, column=9).value is not None or 
           hoja.cell(row=next_row, column=10).value is not None): 
        next_row += 1  # Aumentar la fila hasta encontrar una fila vacía en las tres columnas

    # Ingresar los valores en la siguiente fila vacía
    hoja.cell(row=next_row, column=8, value=cojinete)
    hoja.cell(row=next_row, column=9, value=Temperatura)
    hoja.cell(row=next_row, column=10, value=fecha)

    # Guardar los cambios en el archivo Excel
    book.save(nombre_archivo)
    messagebox.showinfo("Éxito", "Revisión térmica agregada exitosamente.")
    


def actualizar_combobox_archivos():
    carpeta_actual = os.getcwd()
    archivos_excel.clear()
    for archivo in os.listdir(carpeta_actual):
        if archivo.endswith('.xlsx'):
            archivos_excel.append(archivo)

    archivo_combobox['values'] = archivos_excel
    if archivos_excel:
        archivo_combobox.current(0)
        actualizar_combobox_hojas()

def actualizar_combobox_hojas(event=None):
    archivo_seleccionado = archivo_combobox.get()
    if archivo_seleccionado:
        book = load_workbook(archivo_seleccionado)
        hojas = book.sheetnames
        hoja_combobox['values'] = hojas
        hoja_combobox.current(0)

def crear_nuevo_excel():
    nombre_archivo = simpledialog.askstring("Nuevo Excel", "Ingrese el nombre del nuevo archivo de Excel (con .xlsx):")
    if nombre_archivo:
        if not nombre_archivo.endswith('.xlsx'):
            nombre_archivo += '.xlsx'
        crear_excel(nombre_archivo)

def ver_contenido_hoja():
    nombre_archivo = archivo_combobox.get()
    hoja_seleccionada_nombre = hoja_combobox.get()  # Obtener la hoja seleccionada
    
    
    # Intentar abrir el archivo de Excel
    try:
        # Cargar el libro de Excel
        book = load_workbook(nombre_archivo)
    except FileNotFoundError:
        messagebox.showerror("Error", f"No se pudo encontrar el archivo '{nombre_archivo}'")
        return
    
         # Obtener la hoja seleccionada como objeto Worksheet
    hoja_seleccionada = book[hoja_seleccionada_nombre]


    # Crear una nueva ventana emergente que ocupe toda la pantalla
    ventana = Toplevel()
    ventana.title(f"Contenido de la hoja '{hoja_seleccionada.title}'")

    screen_width = ventana.winfo_screenwidth()
    screen_height = ventana.winfo_screenheight()
    ventana.geometry(f"{screen_width}x{screen_height}")

    # Frame para contener el Treeview y las barras de desplazamiento
    frame = tk.Frame(ventana)
    frame.pack(fill=tk.BOTH, expand=True)

    # Crear un Treeview para mostrar los datos
    tree = ttk.Treeview(frame, columns=[str(i) for i in range(1, hoja_seleccionada.max_column + 1)], show='headings')

    # Configurar las barras de desplazamiento
    scroll_y = Scrollbar(frame, orient="vertical", command=tree.yview)
    scroll_x = Scrollbar(frame, orient="horizontal", command=tree.xview)
    
    tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

    scroll_y.pack(side=tk.RIGHT, fill="y")
    scroll_x.pack(side=tk.BOTTOM, fill="x")
    
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    # Configurar encabezados del Treeview
    for col in range(1, hoja_seleccionada.max_column + 1):
        tree.heading(str(col), text=f'Columna {col}')
        tree.column(str(col), width=100, anchor='center')  # Ancho de cada columna ajustado y alineado al centro

    # Insertar los datos de la hoja en el Treeview
    for row in range(1, hoja_seleccionada.max_row + 1):
        fila = []
        for col in range(1, hoja_seleccionada.max_column + 1):
            valor = hoja_seleccionada.cell(row=row, column=col).value
            fila.append(valor if valor is not None else "")
        tree.insert('', 'end', values=fila)

    # Botón para cerrar la ventana emergente
    cerrar_btn = tk.Button(ventana, text="Cerrar", command=ventana.destroy)
    cerrar_btn.pack(pady=10)

# Función para abrir el archivo Excel seleccionado en la aplicación predeterminada
def abrir_excel():
    archivo_seleccionado = archivo_combobox.get()  # Obtener el archivo seleccionado
    hoja_seleccionada = hoja_combobox.get()  # Obtener la hoja seleccionada

    if archivo_seleccionado and hoja_seleccionada:
        # Ruta del archivo
        ruta_archivo = os.path.join(os.getcwd(), archivo_seleccionado)
        
        # Imprimir la ruta para depuración
        print(f"Intentando abrir el archivo en: {ruta_archivo}")
        
        try:
            # Verificar si el archivo existe
            if os.path.isfile(ruta_archivo):
                # Intentar abrir el archivo con la aplicación predeterminada del sistema operativo
                os.startfile(ruta_archivo)  # Esto abrirá el archivo con la aplicación predeterminada de Windows (Excel)
                
                messagebox.showinfo("Archivo abierto", f"Se ha abierto el archivo '{archivo_seleccionado}' en la hoja '{hoja_seleccionada}'.")
            else:
                raise FileNotFoundError(f"No se pudo encontrar el archivo en la ruta: {ruta_archivo}")
        
        except FileNotFoundError as fnf_error:
            messagebox.showerror("Error", str(fnf_error))
    else:
        messagebox.showwarning("Selección incompleta", "Por favor, selecciona tanto un archivo como una hoja.")
        

def montrar_creditos():
    # Crear una nueva ventana secundaria
    ventana_secundaria = Toplevel(root)
    ventana_secundaria.title("Ventana Secundaria")
    ventana_secundaria.geometry("1000x450")

    # Crear varias etiquetas, una debajo de otra, alineadas a la derecha
    oraciones = [
        " -----------------------------------------------------Información importante sobre el programa---------------------------------------------------",
        "Guardar Motor: Guarda los valores de Area, Codigo, etc, de los campos de entrada en la hoja de excel en el que se está trabajando.",
        "Editar Motor: Muestra los motores que hay en la hoja y al escirbir el nombre del motor permite editar cualquiera de los campos de dicho motor.",
        "Agregar mantenimiento: Solicita nombre del técnico, el tipo de mantenimiento y fecha para ir registrando los mantenimientos.",
        "Crear Nueva Hoja: Permite crear una nueva hoja en el excel en el que se está trabajando.",
        "Crear Nuevo Archivo Excel: Permite crear nuevos Excel dentro de la carpeta del programa",
        "Ver Contenido Hoja: Abre una ventana secundaria que muestra el contenido de la hoja del excel en el que se esta trabajando.",
        "Abrir archivo Excel: Abre el archivo de excel seleccionado.",
        "Revisión térmica: Permite ingresar revisiones térmicas de los cojinetes en el motor.",
        "--------------------------------------------------------------------------------------------------------------------------------------------------",
        "Creado por Ing. Kevin Alarcón",
        "Contacto 3303-3211"
    ]
    
    for texto in oraciones:
        label = tk.Label(ventana_secundaria, text=texto, anchor="w", width=150)
        label.pack(anchor='w', padx=10, pady=5)

        

# Función para cambiar el foco entre campos con Enter
def cambiar_foco(event, siguiente_widget):
    siguiente_widget.focus_set()

# Crear la ventana principal
root = Tk()
root.title("Gestión de Motores")
root.geometry("800x800")  # Ajustar tamaño de la ventana a 500x450 para incluir botón de mantenimiento

# Estilos para mejorar la apariencia
font = ('Arial', 10)
button_font = ('Arial', 9, 'bold')

# Cargar y mostrar la imagen en la parte superior
ruta_imagen = "GrandBayGuatemala.png"
image = Image.open(ruta_imagen)
image_tk = ImageTk.PhotoImage(image)
label_imagen = tk.Label(root, image=image_tk)
label_imagen.pack(side="top")  # Aparece en la parte superior

# Crear un combobox para seleccionar el archivo Excel
Label(root, text="Seleccionar Archivo:", font=font).pack(pady=5)
archivo_combobox = ttk.Combobox(root, state="readonly", font=font)
archivo_combobox.pack(pady=5)
archivo_combobox.bind("<<ComboboxSelected>>", actualizar_combobox_hojas)

# Crear un combobox para seleccionar la hoja
Label(root, text="Seleccionar Hoja:", font=font).pack(pady=5)
hoja_combobox = ttk.Combobox(root, state="readonly", font=font)
hoja_combobox.pack(pady=5)

# Crear un Frame para los widgets con grid
grid_frame = Frame(root)
grid_frame.pack(pady=10)


# Campos para guardar el motor
AREA_label = Label(grid_frame, text="AREA")
AREA_label.grid(row=2, column=0, padx=10, pady=5, sticky=W)

Area_entry = Entry(grid_frame)
Area_entry.grid(row=2, column=1, padx=10, pady=5)

CODIGO_label = Label(grid_frame, text="CODIGO")
CODIGO_label.grid(row=3, column=0, padx=10, pady=5, sticky=W)

CODIGO_entry = Entry(grid_frame)
CODIGO_entry.grid(row=3, column=1, padx=10, pady=5)

TIPO_label = Label(grid_frame, text="TIPO")
TIPO_label.grid(row=4, column=0, padx=10, pady=5, sticky=W)

TIPO_entry = Entry(grid_frame)
TIPO_entry.grid(row=4, column=1, padx=10, pady=5)

DATOS_label = Label(grid_frame, text="DATOS")
DATOS_label.grid(row=5, column=0, padx=10, pady=5, sticky=W)

DATOS_entry = Entry(grid_frame)
DATOS_entry.grid(row=5, column=1, padx=10, pady=5)

MARCA_label = Label(grid_frame, text="MARCA")
MARCA_label.grid(row=6, column=0, padx=10, pady=5, sticky=W)

MARCA_entry = Entry(grid_frame)
MARCA_entry.grid(row=6, column=1, padx=10, pady=5)

FUNCION_label = Label(grid_frame, text="FUNCION")
FUNCION_label.grid(row=7, column=0, padx=10, pady=5, sticky=W)

FUNCION_entry = Entry(grid_frame)
FUNCION_entry.grid(row=7, column=1, padx=10, pady=5)

POTENCIA_label = Label(grid_frame, text="POTENCIA (Kw)")
POTENCIA_label.grid(row=8, column=0, padx=10, pady=5, sticky=W)

POTENCIA_entry = Entry(grid_frame)
POTENCIA_entry.grid(row=8, column=1, padx=10, pady=5)

VOLTAJE_label = Label(grid_frame, text="VOLTAJE (V)")
VOLTAJE_label.grid(row=9, column=0, padx=10, pady=5, sticky=W)

VOLTAJE_entry = Entry(grid_frame)
VOLTAJE_entry.grid(row=9, column=1, padx=10, pady=5)

VOLTAJE_FUENTE_label = Label(grid_frame, text="VOLTAJE FUENTE (V)")
VOLTAJE_FUENTE_label.grid(row=2, column=3, padx=10, pady=5, sticky=W)

VOLTAJE_FUENTE_entry = Entry(grid_frame)
VOLTAJE_FUENTE_entry.grid(row=2, column=4, padx=10, pady=5)

AMPERAJE_label = Label(grid_frame, text="AMPERAJE (A)")
AMPERAJE_label.grid(row=3, column=3, padx=10, pady=5, sticky=W)

AMPERAJE_entry = Entry(grid_frame)
AMPERAJE_entry.grid(row=3, column=4, padx=10, pady=5)

COJINETE_FRONTAL_label = Label(grid_frame, text="COJINETE FRONTAL")
COJINETE_FRONTAL_label.grid(row=4, column=3, padx=10, pady=5, sticky=W)

COJINETE_FRONTAL_entry = Entry(grid_frame)
COJINETE_FRONTAL_entry.grid(row=4, column=4, padx=10, pady=5)

COJINETE_TRASERO_label = Label(grid_frame, text="COJINETE TRASERO")
COJINETE_TRASERO_label.grid(row=5, column=3, padx=10, pady=5, sticky=W)

COJINETE_TRASERO_entry = Entry(grid_frame)
COJINETE_TRASERO_entry.grid(row=5, column=4, padx=10, pady=5)

MANTENIMIENTO_label = Label(grid_frame, text="MANTENIMIENTO")
MANTENIMIENTO_label.grid(row=6, column=3, padx=10, pady=5, sticky=W)

MANTENIMIENTO_entry = Entry(grid_frame)
MANTENIMIENTO_entry.grid(row=6, column=4, padx=10, pady=5)

RPM_label = Label(grid_frame, text="RPM")
RPM_label.grid(row=7, column=3, padx=10, pady=5, sticky=W)

RPM_entry = Entry(grid_frame)
RPM_entry.grid(row=7, column=4, padx=10, pady=5)

HERTZ_label = Label(grid_frame, text="HERTZ")
HERTZ_label.grid(row=8, column=3, padx=10, pady=5, sticky=W)

HERTZ_entry = Entry(grid_frame)
HERTZ_entry.grid(row=8, column=4, padx=10, pady=5)

OTROS_label = Label(grid_frame, text="OTROS")
OTROS_label.grid(row=9, column=3, padx=10, pady=5, sticky=W)

OTROS_entry = Entry(grid_frame)
OTROS_entry.grid(row=9, column=4, padx=10, pady=5)


# Enlazamos el campo 'Enter' para pasar entre campos
Area_entry.bind('<Return>', lambda event: cambiar_foco(event, CODIGO_entry))
CODIGO_entry.bind('<Return>', lambda event: cambiar_foco(event, TIPO_entry))
TIPO_entry.bind('<Return>', lambda event: cambiar_foco(event, DATOS_entry))
DATOS_entry.bind('<Return>', lambda event: cambiar_foco(event, MARCA_entry))
MARCA_entry.bind('<Return>', lambda event: cambiar_foco(event, FUNCION_entry))
FUNCION_entry.bind('<Return>', lambda event: cambiar_foco(event, POTENCIA_entry))
POTENCIA_entry.bind('<Return>', lambda event: cambiar_foco(event, VOLTAJE_entry))
VOLTAJE_entry.bind('<Return>', lambda event: cambiar_foco(event, VOLTAJE_FUENTE_entry))
VOLTAJE_FUENTE_entry.bind('<Return>', lambda event: cambiar_foco(event, AMPERAJE_entry))
AMPERAJE_entry.bind('<Return>', lambda event: cambiar_foco(event, COJINETE_FRONTAL_entry))
COJINETE_FRONTAL_entry.bind('<Return>', lambda event: cambiar_foco(event, COJINETE_TRASERO_entry))
COJINETE_TRASERO_entry.bind('<Return>', lambda event: cambiar_foco(event, MANTENIMIENTO_entry))
MANTENIMIENTO_entry.bind('<Return>', lambda event: cambiar_foco(event, RPM_entry))
RPM_entry.bind('<Return>', lambda event: cambiar_foco(event, HERTZ_entry))
HERTZ_entry.bind('<Return>', lambda event: cambiar_foco(event, OTROS_entry))
OTROS_entry.bind('<Return>', lambda event: cambiar_foco(event, Area_entry))


# Botones
# Crear un Frame para los botones
frame_botones = Frame(root)
frame_botones.pack(pady=10)

# Botones de acción
guardar_button = Button(frame_botones, text="           Guardar Motor           ", font=button_font, command=guardar_nuevo_motor)
guardar_button.grid(row=0, column=0, padx=5, pady=5)

editar_button = Button(frame_botones, text="       Editar Motor      ", font=button_font, command=editar_motor)
editar_button.grid(row=0, column=1, padx=5, pady=5)

mantenimiento_button = Button(frame_botones, text="Agregar Mantenimiento", font=button_font, command=agregar_mantenimiento)
mantenimiento_button.grid(row=0, column=2, padx=5, pady=5)

# Crear un botón para crear nuevo archivo Excel
crear_excel_button = Button(frame_botones, text="Crear Nuevo Archivo Excel", font=button_font, command=crear_nuevo_excel)
crear_excel_button.grid(row=1, column=0, padx=10, pady=5)

# Crear un botón para crear nuevo archivo Excel
crear_hoja_button = Button(frame_botones, text="Crear Nueva Hoja", font=button_font, command=lambda: crear_hoja(archivo_combobox.get()))
crear_hoja_button.grid(row=0, column=3, padx=10, pady=5)

# Crear botón para ver el contenido de la hoja
ver_contenido_button = Button(frame_botones, text="Ver Contenido Hoja", font=button_font, command=ver_contenido_hoja)
ver_contenido_button.grid(row=1, column=1, padx=10, pady=5)

# Crear el botón que abrirá el archivo de Excel
abrir_excel_button = Button(frame_botones, text="    Abrir archivo Excel    ", font=button_font, command=abrir_excel)
abrir_excel_button.grid(row=1, column=2, padx=10, pady=10)

# Crear un botón para ingresar revisiones térmicas
Mantenimiento_button = Button(frame_botones, text=" Revisión térmica ", font=button_font, command=agregar_revision_termica)
Mantenimiento_button.grid(row=1, column=3, padx=10, pady=5)

# Crear un botón para créditos
Mantenimiento_button = Button(frame_botones, text="Créditos", font=button_font, command=montrar_creditos)
Mantenimiento_button.grid(row=2, column=3, padx=10, pady=5)

# Actualizamos comboboxes al inicio
actualizar_combobox_archivos()

root.mainloop()